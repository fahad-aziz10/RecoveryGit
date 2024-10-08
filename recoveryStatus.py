import time, sys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities


from openpyxl import load_workbook
caps = DesiredCapabilities().EDGE
# caps["pageLoadStrategy"] = "normal"  #  complete
caps["pageLoadStrategy"] = "eager"  #  interactive
#caps["pageLoadStrategy"] = "none"   #  undefined
# driver = webdriver.Opera( executable_path=r'C:\\Drivers\\edgedriver_win64.exe')
# desired_capabilities=caps,


############################################
if len(sys.argv) == 5:
    parentPath = str(sys.argv[1])+'\\'
    fileToProcess = str(sys.argv[2])+'.xlsx'
    refNoCol = int(sys.argv[3])
    remarksCol = int(sys.argv[4])
else:
######## Section for Manual Input ##########
    parentPath = input('Enter folder path (e.g.,D:\\Lesco\\Recovery\\01-23) :')+'\\'
    fileToProcess = input('Enter file name (e.g.,WorkingBookDT25-10k) :')+'.xlsx'
    refNoCol = int(input('Enter reference number column number (e.g.,3) :'))
    remarksCol = int(input('Enter remarks number (e.g.,3) :'))
isRefComplete = True
batchCol = 2
subDiv  =   "11216"
############################################
driver = webdriver.Chrome(desired_capabilities=caps,executable_path='C:/Drivers/edgedriver_win64/msedgedriver.exe')
wb = load_workbook(filename = parentPath+fileToProcess)
    
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print("Testing of "+ sheet +"started from row 2 to "+str(ws.max_row+1))
    for row in range(2,ws.max_row+1): 
        # if row%100==0:
                # print('Saving Workbook')
                # wb.save(filename = parentPath+sheet +'_'+ str(row)+'_'+fileToProcess)
        if isRefComplete:
            ref     =   str(ws.cell(row,refNoCol).value) 
            batch   =   ref[:len(ref)-12]#str(ws.cell(row,batchCol).value)
            subDiv  =   ref[-12:-7]
            refNo   =   ref[-7:]
        else:
            batch   =   str(ws.cell(row,batchCol).value)
            refNo   =   str(ws.cell(row,refNoCol).value)
        print("Batch : "+str(batch)+"\tSubDiv : "+subDiv+"\tRef No. : "+refNo)
        print(str(ws.cell(row,remarksCol).value))
        if str(ws.cell(row,remarksCol).value) != 'None' :
            print("Remarks already present")
            continue
        
        link = "http://www.lesco.gov.pk:36269/Modules/CustomerBillN/CheckBill.asp"
        # link = "http://www.lesco.gov.pk:36269/Modules/CustomerBill/CheckBill.asp"
        # if batch == '24' or batch == '44' or batch == '46' or batch == '36':
        #     link = "http://www.lesco.gov.pk/Customer_Reg/AccountStatusMDI.aspx?nBatchNo="+batch+"&nSubDiv="+subDiv+"&nRefNo="+refNo+"&strRU=U"
        # else:
        #     link = "http://www.lesco.gov.pk/Customer_Reg/AccountStatus.aspx?nBatchNo="+batch+"&nSubDiv="+subDiv+"&nRefNo="+refNo+"&strRU=U"
        try:
            driver.get(link)
            WebDriverWait(driver, timeout=3).until(ec.visibility_of_element_located((By.XPATH , '/html/body/div[1]/div/div[2]/form[1]/center/div/table/tbody/tr[1]/td/input[1]')))
            driver.find_element(By.XPATH,'/html/body/div[1]/div/div[2]/form[1]/center/div/table/tbody/tr[1]/td/input[1]').send_keys(batch)
            driver.find_element(By.XPATH,'/html/body/div[1]/div/div[2]/form[1]/center/div/table/tbody/tr[1]/td/input[2]').send_keys(subDiv)
            driver.find_element(By.XPATH,'/html/body/div[1]/div/div[2]/form[1]/center/div/table/tbody/tr[1]/td/input[3]').send_keys(refNo)
            driver.find_element(By.XPATH,'/html/body/div[1]/div/div[2]/form[1]/center/div/table/tbody/tr[2]/td/input[1]').click()

            WebDriverWait(driver, timeout=3).until(ec.visibility_of_element_located((By.XPATH , '/html/body/div[1]/div/div[2]/table/tbody/tr[9]/td/form/button')))
            driver.find_element(By.XPATH,'/html/body/div[1]/div/div[2]/table/tbody/tr[9]/td/form/button').click()

            WebDriverWait(driver, timeout=3).until(ec.visibility_of_element_located((By.XPATH , '/html/body/div[1]/div[2]/div/div[4]/div[1]/strong')))
            output_txt = 'P='+driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div[4]/div[1]/strong').text
            if output_txt == 'P=0':
                print("No Payment")
                continue
            output_txt += '  DT '+driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div[4]/div[2]/strong').text
            output_txt += '  IN '+driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div[4]/div[3]/strong').text
            ws.cell(row,remarksCol).value = output_txt
            print(output_txt)
        except:
            print('No Data Found')

    try:
        print('Completed Sheet '+sheet+' Saving Workbook')
        wb.save(filename = parentPath+'complete_'+sheet+'_'+fileToProcess)
    except:
        pass
print('Completed File Saving Workbook')
try:
    wb.save(filename = parentPath+'complete_'+fileToProcess)
    wb.close()
    driver.quit()
except:
    retry ='r'
    while retry =='r':
        wb.save(filename = parentPath+'complete_'+fileToProcess)
        wb.close()
        driver.quit()
        retry = input("Enter 'r' to retry, any other key to exit: ")



